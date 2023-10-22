from aiohttp import web
from sqlalchemy.orm import Session
from db.models import User
from typing import Callable, Optional, Any
from functools import wraps
from email_validator import validate_email, EmailNotValidError
import enum
from api.schemas import BookCreate
import uuid
import config
import openpyxl
import os


class ValidatorType(enum.Enum):
    create_user = 'create_user'
    create_book = 'create_book'
    download_book = 'download_book'
    create_author = 'create_author'
    denied_books = 'denied_books'


async def upload_file(book_file: Any, path: str = f'{config.UPLOADS_PATH}/books') -> None:
    file_name = str(uuid.uuid4())
    upload_dir = path
    old_end_file_name = book_file.filename.split(".")[-1]
    if old_end_file_name:
        file_name += f'.{old_end_file_name}'

    os.makedirs(upload_dir, exist_ok=True)
    file_path = os.path.join(upload_dir, file_name)
    with open(file_path, 'wb') as f:
        while True:
            chunk = await book_file.read_chunk()
            if not chunk:
                break
            f.write(chunk)
    return file_name


def fullfil_data_info(file_path: str) -> list[tuple[str, str]]:
    workbook = openpyxl.load_workbook(file_path)
    # Access the 'author' sheet by name
    author_sheet = workbook['author']

    # Access the 'name' sheet by name
    name_sheet = workbook['name']

    author_data = []
    name_data = []

    # Read data from the 'author' sheet
    for row in author_sheet.iter_rows(values_only=True):
        if row[1] == 'name':
            continue
        author_data.append(row[1])  # Assuming the data you want is in the second column (index 1)

    # Read data from the 'name' sheet
    for row in name_sheet.iter_rows(values_only=True):
        if row[1] == 'name':
            continue
        name_data.append(row[1])  # Assuming the data you want is in the second column (index 1)
    book_info = []
    for i in range(min(len(name_data), len(author_data))):
        book_info.append((name_data[i], author_data[i]))
    return book_info


async def validator(request: web.Request, validator_type: ValidatorType) -> Any:
    if validator_type == ValidatorType.create_user:
        data = await request.json()
        email = data.get('email')
        code = data.get('code')

        try:
            emailinfo = validate_email(email, check_deliverability=False)
            email = emailinfo.normalized
        except EmailNotValidError as e:
            print(str(e))
        if code != 123456:
            raise web.HTTPBadRequest('code is wrong')
        return {
            'email': email
        }
    if validator_type == ValidatorType.create_book:
        reader = await request.multipart()
        # Initialize variables to store form fields
        data = {
            'name': None,
            'author_id': None,
        }
        while True:
            field = await reader.next()
            if field is None:
                break
            if field.filename:
                data['file_name'] = await upload_file(field)
            else:
                field_name = field.name
                field_value = await field.text()
                if field_name in BookCreate.model_fields:
                    try:
                        field_value = BookCreate.model_fields[field_name]._attributes_set['annotation'](field_value)
                        data[field_name] = field_value
                    except ValueError as e:
                        raise e
                else:
                    raise web.HTTPBadRequest(f'Validation error: Wrong field_name: {field_name}')
        if not data['name'] or not data['author_id']:
            raise web.HTTPBadRequest('Validation error: Please provide a book_file, book_name, and author_name fields.')
        return data

    if validator_type == ValidatorType.create_author:
        data = await request.json()
        author_name = data.get('author_name')
        return {
            'author_name': author_name
        }

    if validator_type == ValidatorType.download_book:
        try:
            book_id = int(request.match_info['id'])
        except Exception:
            raise ValueError('wrong book_id')
        return {'book_id': book_id}

    if validator_type == ValidatorType.denied_books:
        reader = await request.multipart()
        data = {}
        while True:
            field = await reader.next()
            if field is None:
                break
            if field.filename:
                path = f'{config.UPLOADS_PATH}/denied_requests'
                data['file_name'] = await upload_file(field, path)

        if not data['file_name']:  # TODO add better validation
            raise web.HTTPBadRequest('Validation error: need to get xlsx file')

        file_path = f'{config.UPLOADS_PATH}/denied_requests/{data["file_name"]}'
        data['books_info'] = fullfil_data_info(file_path)
        return data

    return {}


# TODO annotate *args, **kwargs by TypeVar, ParamSpec
def login_required(
    admin_require: bool = False, validator_type: Optional[ValidatorType] = None
) -> Callable:
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        async def wrapped(request, *args, **kwargs) -> None:
            # TODO add tokens in header
            # authorized = False
            # token = request.headers.get('X-G-Token')

            data = await validator(request, validator_type)
            db: Session = request['session']
            user_query = db.query(User).filter(~User.registration_completed_at.is_(None))
            if admin_require:
                user_query = user_query.filter(User.is_admin.is_(True))
            user = user_query.first()
            if not user:
                raise web.HTTPBadRequest(text='check or refresh tokens, cannot find user')
            return await func(request, db, user, *args, **kwargs, **data)
        return wrapped

    return decorator


def login_not_required(admin_require: bool = False, validator_type: Optional[ValidatorType] = None) -> Callable:
    def decorator(func: Callable) -> Callable:
        @wraps(func)
        async def wrapped(request, *args, **kwargs) -> None:
            # authorized = False
            # token = request.headers.get('X-G-Token')
            db: Session = request['session']
            data = await validator(request, validator_type)
            return await func(request, db, *args, **kwargs, **data,)
        return wrapped

    return decorator
